# Carson Chadwick
# Section 04
# This program takes excel data and organizes it into a new workbook based on the classes the teachers teach and information about each class

# Import Libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Load workbook
myWorkbook = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
currWS = myWorkbook.active

# Loop to itterate through each value in the Grades sheet, creates sheets for each class, adds and formats headers for each seet, and loads student data
for objRow in currWS.iter_rows(min_row =2) :
    if objRow[0].value not in myWorkbook.sheetnames :
        myWorkbook.create_sheet(objRow[0].value)
        activeWS = myWorkbook[objRow[0].value]
        activeWS["A1"] = "Last Name"
        activeWS["A1"].font = Font(bold=True)
        activeWS.column_dimensions["A"].width = len(myWorkbook[objRow[0].value]["A1"].value) + 5
        activeWS["B1"] = "First Name"
        activeWS["B1"].font = Font(bold=True)
        activeWS.column_dimensions["B"].width = len(myWorkbook[objRow[0].value]["B1"].value) + 5
        activeWS["C1"] = "Student ID"
        activeWS["C1"].font = Font(bold=True)
        activeWS.column_dimensions["C"].width = len(myWorkbook[objRow[0].value]["C1"].value) + 5
        activeWS["D1"] = "Grade"
        activeWS["D1"].font = Font(bold=True)
        activeWS.column_dimensions["D"].width = len(myWorkbook[objRow[0].value]["D1"].value) + 5
        activeWS["F1"] = "Summary Statistics"
        activeWS["F1"].font = Font(bold=True)
        activeWS.column_dimensions["F"].width = len(myWorkbook[objRow[0].value]["F1"].value) + 5
        activeWS["G1"] = "Value"
        activeWS["G1"].font = Font(bold=True)
        activeWS.column_dimensions["G"].width = len(myWorkbook[objRow[0].value]["D1"].value) + 5
    lstInfo = objRow[1].value.split("_")
    lstInfo.append(objRow[2].value)
    myWorkbook[objRow[0].value].append(lstInfo)

# For each sheet adjust the header lengths
for sheet in myWorkbook.sheetnames :
    max = myWorkbook[sheet].max_row
    range = "A1:D" + str(max)
    myWorkbook[sheet].auto_filter.ref = (range)
    # In each sheet also add Highest, Lowest, Mean, Median, and Count of grades
    gradeCol = "D2:D" + str(max)
    myWorkbook[sheet]["F2"] = "Highest Grade" 
    myWorkbook[sheet]["G2"] = f"=MAX({gradeCol})"
    myWorkbook[sheet]["F3"] = "Lowest Grade"
    myWorkbook[sheet]["G3"] = f"=MIN({gradeCol})"
    myWorkbook[sheet]["F4"] = "Mean Grade"
    myWorkbook[sheet]["G4"] = f"=AVERAGE({gradeCol})"
    myWorkbook[sheet]["F5"] = "Median Grade"
    myWorkbook[sheet]["G5"] = f"=MEDIAN({gradeCol})"
    myWorkbook[sheet]["F6"] = "Number of Students"
    myWorkbook[sheet]["G6"] = f"=COUNT({gradeCol})"

# Remove the grades sheet, save, and close
myWorkbook.remove(myWorkbook["Grades"])
myWorkbook.save(filename="formatted_grades.xlsx")
myWorkbook.close()