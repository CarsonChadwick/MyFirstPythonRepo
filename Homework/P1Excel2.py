# Carson Chadwick, Bodhe Melver, John Gibson, Mathew Diamond, Ryan Hagerty, Caleb Stephenson
# Section 04
# This program takes excel data and organizes it into a new workbook based on the classes the teachers teach and information about each class

# Import Libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Load workbook
myWorkbook = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
currWS = myWorkbook.active
# Create new Workbook to save data
newBook = Workbook()
newBook.remove(newBook["Sheet"])
# Loop to itterate through each value in the Grades sheet, creates sheets for each class, adds and formats headers for each seet, and loads student data
for objRow in currWS.iter_rows(min_row =2) :
    if objRow[0].value not in newBook.sheetnames :
        newBook.create_sheet(objRow[0].value)
        activeWS = newBook[objRow[0].value]
        activeWS["A1"] = "Last Name"
        activeWS["A1"].font = Font(bold=True)
        activeWS.column_dimensions["A"].width = len(newBook[objRow[0].value]["A1"].value) + 5
        activeWS["B1"] = "First Name"
        activeWS["B1"].font = Font(bold=True)
        activeWS.column_dimensions["B"].width = len(newBook[objRow[0].value]["B1"].value) + 5
        activeWS["C1"] = "Student ID"
        activeWS["C1"].font = Font(bold=True)
        activeWS.column_dimensions["C"].width = len(newBook[objRow[0].value]["C1"].value) + 5
        activeWS["D1"] = "Grade"
        activeWS["D1"].font = Font(bold=True)
        activeWS.column_dimensions["D"].width = len(newBook[objRow[0].value]["D1"].value) + 5
        activeWS["F1"] = "Summary Statistics"
        activeWS["F1"].font = Font(bold=True)
        activeWS.column_dimensions["F"].width = len(newBook[objRow[0].value]["F1"].value) + 5
        activeWS["G1"] = "Value"
        activeWS["G1"].font = Font(bold=True)
        activeWS.column_dimensions["G"].width = len(newBook[objRow[0].value]["D1"].value) + 5
    # Splits up the data in the second column of grades and adds it to list
    lstInfo = objRow[1].value.split("_")
    # Add the grades data to the list
    lstInfo.append(objRow[2].value)
    # Add the list to the new wookbook in the next column.
    newBook[objRow[0].value].append(lstInfo)

# For each sheet adjust the header lengths
for sheet in newBook.sheetnames :
    max = newBook[sheet].max_row
    range = "A1:D" + str(max)
    newBook[sheet].auto_filter.ref = (range)
    # In each sheet also add Highest, Lowest, Mean, Median, and Count of grades
    gradeCol = "D2:D" + str(max)
    newBook[sheet]["F2"] = "Highest Grade" 
    newBook[sheet]["G2"] = f"=MAX({gradeCol})"
    newBook[sheet]["F3"] = "Lowest Grade"
    newBook[sheet]["G3"] = f"=MIN({gradeCol})"
    newBook[sheet]["F4"] = "Mean Grade"
    newBook[sheet]["G4"] = f"=AVERAGE({gradeCol})"
    newBook[sheet]["F5"] = "Median Grade"
    newBook[sheet]["G5"] = f"=MEDIAN({gradeCol})"
    newBook[sheet]["F6"] = "Number of Students"
    newBook[sheet]["G6"] = f"=COUNT({gradeCol})"

# Remove the grades sheet, save, and close
newBook.save(filename="formatted_grades.xlsx")
newBook.close()
myWorkbook.close()