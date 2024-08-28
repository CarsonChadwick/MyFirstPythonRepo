import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

myWorkbook = openpyxl.load_workbook('StudentData.xlsx')

currWS = myWorkbook.active
currWS.auto_filter.ref = "A1:F11"
f1 = Font(bold=True)
f2 = Font(color='FF0000')
f3 = Font(color='00FF00')
f4 = Font(color='800080')
f5 = Font(color='808080')

currWS.column_dimensions["A"].auto_size = True
currWS.column_dimensions["B"].auto_size = True
currWS.column_dimensions["C"].auto_size = True
currWS.column_dimensions["D"].auto_size = True
currWS.column_dimensions["E"].auto_size = True
currWS.column_dimensions["F"].auto_size = True 

currWS["A1"].font = f1
currWS["B1"].font = f1
currWS["C1"].font = f1
currWS["D1"].font = f1
currWS["E1"].font = f1
currWS["F1"].font = f1

# Display the CLASS data using FF0000 for FR, 00FF00 for SO, 800080 for JR, and 808080 for SR.
cols = ["A", "B", "C", "D", "E", "F"]

for cell in currWS['E'] :
    if cell.value == "FR":
        currWS[cell.coordinate].font = f2
    elif cell.value == "SO":
        currWS[cell.coordinate].font = f3
    elif cell.value == "JR":
        currWS[cell.coordinate].font = f4
    elif cell.value =="SR" :
        currWS[cell.coordinate].font = f5
    
#Center GPA, CLASS, and AGE in columns D, E, F
for cell in currWS['D'] :
    currWS[cell.coordinate].alignment = Alignment(horizontal="center")
for cell in currWS['E'] :
    currWS[cell.coordinate].alignment = Alignment(horizontal="center")
for cell in currWS['F'] :
    currWS[cell.coordinate].alignment = Alignment(horizontal="center")

#Calculate GPA average
currWS["D12"] = '=AVERAGE(D2:D11)'

#Max AGE
currWS["H2"] = 'Max Age'
currWS["I2"] = '=MAX(F2:F11)'

currWS["H3"] = 'Min Age'
currWS["I3"] = '=MIN(F2:F11)'

currWS["A15"] = '# of Srs'
currWS["B15"] = '=COUNTIF(E2:E11, "SR")'
myWorkbook.save(filename="StudentData.xlsx")
#display all the data
for row in currWS.iter_rows(values_only=True) :
    print(row)
myWorkbook.close()


