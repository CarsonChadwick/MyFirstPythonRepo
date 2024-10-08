import os
import platform
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#Not required but cool
def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')
#Not required but cool
clear_screen()
#Open Excel File for reading
myWorkbook = openpyxl.load_workbook("sampleFood.xlsx")
#Set active sheet
currSheet = myWorkbook.active
#Make a new workbook for writing
newBook = Workbook()
#Remove current worksheet
newBook.remove(newBook["Sheet"])
#Read the file
#You need to know your data!
#min_row : Which row to start on
#max_col : Number of coluns to read
#max_row : Number of rows to read
for row in currSheet.iter_rows(min_row = 2, max_col = 1, max_row = 9):
    #Read first column value for current row
    #data is returned as a list
    data =row[0].value
    #Split the data based upon the underscore symbol
    data = data.split(',')
    #This creates a list with 2 positions
    #[0] is the category
    #[1] is the description
    #[2] is the calorie info
    #Store the split data into variables
    category = data[0]
    description = data[1]
    calories = int(data[2]) #Otherwise it thinks it is str
    #next open the respective sheet and start copying over the data
    #select algebra and paste the algebra info
    if category not in newBook.sheetnames:
        newSheet = newBook.create_sheet(title=category)
        # Add info to the new sheet
        newSheet.append(["Description", "Calories"])
        #adds all the filters to the headers of each sheet
        newSheet.auto_filter.ref = "a1:b1"
    else:
        newSheet = newBook[category]
    #adds the data to the new sheet
    newSheet.append([description, calories])
#adds formulas table to each sheet
for sheet in newBook.sheetnames:
    currentSheet = newBook[sheet]
    #Keep track of the row below the max
    iMaxRow = currentSheet.max_row + 1
    #Keep track of the column to the right of the last column
    iMaxCol = currentSheet.max_column + 1
    #Access a column based upon the position
    #get_column_letter takes an integer and returns a character
    #For example, 1 returns A, 2 returns B, etc.
    currentSheet[get_column_letter(iMaxCol) + "1"] = "Total Calories"
    currentSheet[get_column_letter(iMaxCol) + "2"] = "=SUM(B2:B" + str(currentSheet.max_row) + ")"
    #Process all of the cells on the first row (i.e start row, end row)
    for cell in currentSheet["1:1"]:
        #Change the cell to be bold
        #This makes each cell in the first row bold
        cell.font = Font(bold=True)
        #If the cell has a value
        if cell.value is not None:
            #add 5 to each column width for the cell in the worksheet
            currentSheet.column_dimensions[cell.column_letter].width = len(str(cell.value)) + 5
#new Excel file name
output_file = "format_food.xlsx"
#write workbook in memory to file
newBook.save(filename=output_file)
#close the Excel file
myWorkbook.close()
#Not required but cool
#open the file
os.system(f"start excel {output_file}")