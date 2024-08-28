from openpyxl import Workbook

myWorkbook = Workbook()
wsCustomer = myWorkbook.create_sheet("Customers")
myWorkbook.create_sheet("Food Items")
myWorkbook.create_sheet("Orders")

myWorkbook.active = wsCustomer
wsCustomer["B1"] = "Little Debbie Swiss Cakes"

myWorkbook["Food Items"]["A1"] = "SPAM"
print(myWorkbook["Food Items"]["A1"].value)
myWorkbook.save(filename="test.xlsx")
myWorkbook.close()