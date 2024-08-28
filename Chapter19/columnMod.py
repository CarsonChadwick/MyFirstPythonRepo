from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

myWorkbook = Workbook()
currSheet = myWorkbook.active

f1 = Font(name='Times New Roman', size=16, bold=True, italic=True, color='2E86C1')
f2 = Font(name='Arial', size=20, underline='single')
fill_1 = PatternFill(start_color='FFFF00', end_color='84BF8E', fill_type='darkTrellis')
border_1 = Border(top=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'), bottom=Side(style="thick"))
align_1 = Alignment(horizontal="center", text_rotation=90, wrap_text=True)

currSheet["A1"] = "First Name"
currSheet["B1"] = "Last Name"
currSheet["C1"] = "Favorite Food"

currSheet["A2"] = "Carmen"
currSheet["B2"] = "Sandiego"
currSheet["C2"] = "Fish"

a1 = currSheet["A1"]
a1.font = f1
a1.fill = fill_1
a1.border = border_1
a1.alignment = align_1

b1 = currSheet["B1"]
b1.font = f2




myWorkbook.save(filename="modifyCol.xlsx")
myWorkbook.close()
