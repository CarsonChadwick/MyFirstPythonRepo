# Carson Chadwick
# Section 04
# This program creates a grade calculator in excel!

# Import Libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill,  Alignment, numbers


f1 = Font(name='Aptos Narrow', size=11, bold=True)
fill_1 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type='darkGray')
align_1 = Alignment(horizontal="right")
numFormat = '0.0'


workBook = Workbook()
currWs = workBook.active
currWs["A1"] = "Assignments (Drop lowest)"
currWs.column_dimensions["A"].width = len(currWs["A1"].value)
currWs["A1"].font = f1
currWs["B1"] = int(input("Enter your average score for your assignments: "))
currWs["C1"] = 0.15
currWs["D1"] = '=B1*C1'
currWs["D1"].number_format = numFormat


currWs["A3"] = "Project 1"
currWs["A3"].font = f1
currWs["B3"] = int(input("Enter your score for Project 1: "))
currWs["C3"] = 0.1
currWs["D3"] = '=B3*C3'
currWs["D3"].number_format = numFormat

currWs["A4"] = "Project 2"
currWs["A4"].font = f1
currWs["B4"] = int(input("Enter your score for Project 2: "))
currWs["C4"] = .1
currWs["D4"] = '=B4*C4'
currWs["D4"].number_format = numFormat

currWs["A5"] = "Project 3"
currWs["A5"].font = f1
currWs["B5"] = int(input("Enter your Score for Project 3: "))
currWs["C5"] = 0.15
currWs["D5"] = '=B5*C5'
currWs["D5"].number_format = numFormat

currWs["A7"] = "Midterm"
currWs["A7"].font = f1
currWs["B7"] = int(input("Enter your score for your midterm: "))
currWs["C7"].fill = fill_1
currWs["D7"].fill = fill_1

currWs["A8"] = "Midterm Exam Credit"
currWs["A8"].font = f1
currWs["B8"] = int(input("Enter how many extra credit points you got on the midterm: "))
currWs["C8"] = 0.2
currWs["D8"] = '=(B7+B8)*C8'
currWs["D8"].number_format = numFormat

currWs["A10"] = "Final"
currWs["A10"].font = f1
currWs["B10"] = int(input("Enter your Final Exam grade: "))
currWs["C10"].fill = fill_1
currWs["D10"].fill = fill_1

currWs["A11"] = "Final Exam Extra Credit"
currWs["A11"].font = f1
currWs["B11"] = int(input("Enter how many points of Extra Credit you got on the Final Exam: "))
currWs["C11"] = 0.3
currWs["D11"] = '=(B10+B11)*C11'
currWs["D11"].number_format = numFormat

currWs["B13"] = "Calculated Grade"
currWs.merge_cells("B13:C13")
currWs["B13"].font = f1
currWs["B13"].fill = fill_1
currWs["B13"].alignment = align_1
currWs["D13"] = "=SUM(D1:D11)"
currWs["D13"].number_format = numFormat

currWs["A15"] = "SONA"
currWs["A15"].font = f1
currWs["B15"] = .25*int(input("Enter how many Sona Credits you did (0-4): "))

currWs["A17"] = "Final Grade without Curve"
currWs.merge_cells('A17:C17')
currWs["A17"].font = f1
currWs["A17"].alignment = align_1
currWs["A17"].fill = fill_1
currWs["D17"] = "=D13+B15"
currWs["D17"].number_format = numFormat

workBook.save(filename="GradeCalc2.xlsx")
workBook.close()










